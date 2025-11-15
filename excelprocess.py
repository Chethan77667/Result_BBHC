import os
import argparse
import logging
import re
from typing import Dict, List, Tuple

import pandas as pd


# Configure logging for CLI usage
logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


# ===== Core helpers (mirrors logic from result_analysis_app.py, no UI) =====

def extract_subject_dict(file_path_or_buffer) -> Dict[str, str]:
    """Read Sheet1 and build a subject-code -> subject-name dictionary."""
    try:
        df_subjects = pd.read_excel(file_path_or_buffer, sheet_name="Sheet1", header=None)
        subjects_dict: Dict[str, str] = {}
        for i in range(len(df_subjects)):
            row = df_subjects.iloc[i]
            subject_code = str(row[3]).strip()
            subject_name = str(row[4]).strip()
            if len(subject_code) == 10 and subject_code.isalnum():
                subjects_dict[subject_code.upper()] = subject_name
        if not subjects_dict:
            raise ValueError("No subject codes found on Sheet1. Expected code in col 3 and name in col 4.")
        return subjects_dict
    except Exception as e:
        logger.error(f"Failed to extract subjects: {e}")
        raise


def extract_usns(df: pd.DataFrame) -> List[Tuple[str, int, int]]:
    """Find each student's USN, the row index of their block, and the column index of the USN label."""
    student_blocks: List[Tuple[str, int, int]] = []
    for i in range(len(df)):
        row = df.iloc[i]
        for col in range(len(row)):
            cell_val = row.iloc[col]
            if isinstance(cell_val, str) and cell_val.strip().upper() == "USN":
                for usn_col in range(col + 1, len(row)):
                    val = row.iloc[usn_col]
                    if isinstance(val, str) and val.strip().upper().startswith("U"):
                        student_blocks.append((val.strip().upper(), i, col))
                        break
                break
    return student_blocks


def get_result_for_usn_subject(
    block: pd.DataFrame,
    input_usn: str,
    subject_code: str,
    subjects_dict: Dict[str, str],
    student_name: str = "Not Found",
):
    subject_code = subject_code.strip().upper()
    for i in range(len(block)):
        row = block.iloc[i]
        if subject_code in row.values:
            col_idx = list(row).index(subject_code)
            try:
                result_status = block.iloc[i + 10, col_idx]
                marks_scored = block.iloc[i + 7, col_idx]
                max_marks = block.iloc[i + 8, col_idx]
            except Exception:
                result_status = ""
                marks_scored = ""
                max_marks = ""

            return {
                "Name": student_name.upper() if student_name else "Not Found",
                "USN": input_usn,
                f"{subject_code} - {subjects_dict[subject_code]}": result_status,
                f"{subject_code}_marks": marks_scored,
                f"{subject_code}_max": max_marks,
            }

    return {
        "Name": "Not Found",
        "USN": input_usn,
        f"{subject_code} - {subjects_dict.get(subject_code, subject_code)}": "N/A",
        f"{subject_code}_marks": "N/A",
        f"{subject_code}_max": "N/A",
    }


def get_gpa_result_total_percent(block: pd.DataFrame, input_usn: str):
    sgpa = cgpa = result_status = "Not Found"
    found_result = False
    for i in range(len(block)):
        row = block.iloc[i]
        for col in range(len(row)):
            val = row.iloc[col]
            if not isinstance(val, str):
                continue
            label = val.strip().upper()
            if label == "SGPA" and sgpa == "Not Found":
                next_val = row.iloc[col + 1] if col + 1 < len(row) else None
                sgpa = next_val if pd.notna(next_val) else "Not Found"
            elif label == "CGPA" and cgpa == "Not Found":
                next_val = row.iloc[col + 1] if col + 1 < len(row) else None
                cgpa = next_val if pd.notna(next_val) else "Not Found"
            elif label.startswith("RESULT:") and not found_result:
                result_status = label.split(":", 1)[-1].strip().upper()
                found_result = True
        if sgpa != "Not Found" and cgpa != "Not Found" and found_result:
            break
    return {"USN": input_usn, "SGPA": sgpa, "CGPA": cgpa, "Result": result_status}


def get_total_and_percentage(block: pd.DataFrame):
    total_marks = max_total = percentage = None

    def parse_numeric(value):
        if pd.isna(value):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        value_str = str(value).strip()
        if value_str == "":
            return None
        # Handle formats like "233.5", "233", "233.5/300", "233-ABSENT"
        match = re.search(r"-?\d+(\.\d+)?", value_str)
        if match:
            try:
                return float(match.group(0))
            except ValueError:
                return None
        return None

    max_total_found = None
    total_found = None

    for i in range(len(block)):
        row = block.iloc[i]
        for col in range(len(row)):
            val = row.iloc[col]
            if not isinstance(val, str):
                continue
            label = val.strip().upper()

            if "MAX" in label and "TOTAL" in label:
                for idx in range(col + 1, len(row)):
                    numeric_val = parse_numeric(row.iloc[idx])
                    if numeric_val is not None:
                        if max_total_found is None or idx > max_total_found[0] or (
                            idx == max_total_found[0] and numeric_val > max_total_found[1]
                        ):
                            max_total_found = (idx, numeric_val)
                continue

            if label == "TOTAL" or (
                "TOTAL" in label and "MAX" not in label and "SUB" not in label
            ):
                for idx in range(col + 1, len(row)):
                    numeric_val = parse_numeric(row.iloc[idx])
                    if numeric_val is not None:
                        if total_found is None or idx > total_found[0] or (
                            idx == total_found[0] and numeric_val > total_found[1]
                        ):
                            total_found = (idx, numeric_val)

    if total_found is not None:
        total_marks = total_found[1]
    if max_total_found is not None:
        max_total = max_total_found[1]

    if total_marks is not None and max_total not in (None, 0):
        percentage = round((float(total_marks) / float(max_total)) * 100, 2)

    return {"Total Marks": total_marks, "Max Marks": max_total, "Percentage": percentage}


def clean_to_int(val):
    if isinstance(val, float) and val.is_integer():
        return int(val)
    return val


def process_excel(input_path: str, mode: str = "combined") -> pd.DataFrame:
    """
    Process the provided Excel file (same layout as used by the app) and
    return the consolidated DataFrame.

    mode:
      - "status": subject columns contain Pass/Fail, etc.
      - "marks": subject columns contain "obtained/max" only
      - "combined": subject columns contain "obtained-status" (e.g., "57-PASS") (default)
    """
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path}")

    df = pd.read_excel(input_path, sheet_name="Sheet2", header=None)
    subjects_dict = extract_subject_dict(input_path)
    student_blocks = extract_usns(df)

    if not student_blocks:
        raise ValueError("No student records found on Sheet2.")

    full_data = []
    for usn, start_idx, usn_col in student_blocks:
        top = max(start_idx - 5, 0)
        bottom = min(start_idx + 12, len(df))
        block = df.iloc[top:bottom]
        name_candidate = None
        if 0 <= start_idx + 1 < len(df):
            name_cell = df.iloc[start_idx + 1, usn_col]
            if isinstance(name_cell, str) and name_cell.strip():
                name_candidate = name_cell.strip()
        student_row = {
            "USN": usn,
            "Name": name_candidate.upper() if name_candidate else "Not Found",
        }

        # Subject-wise
        for code in subjects_dict:
            result = get_result_for_usn_subject(
                block, usn, code, subjects_dict, student_row["Name"]
            )
            marks = result.get(f"{code}_marks", "")
            max_m = result.get(f"{code}_max", "")
            status = result.get(f"{code} - {subjects_dict[code]}", "")
            
            if mode == "marks":
                value = f"{marks}/{max_m}" if marks != "" and max_m != "" else ""
            elif mode == "status":
                value = status
            else:  # mode == "combined" (default)
                if marks != "" and marks != "N/A" and max_m != "" and max_m != "N/A" and status != "" and status != "N/A":
                    value = f"{marks}/{max_m}-{status}"
                elif marks != "" and marks != "N/A" and max_m != "" and max_m != "N/A":
                    value = f"{marks}/{max_m}"
                elif marks != "" and marks != "N/A" and status != "" and status != "N/A":
                    value = f"{marks}-{status}"
                elif marks != "" and marks != "N/A":
                    value = str(marks)
                elif status != "" and status != "N/A":
                    value = status
                else:
                    value = "N/A"
            
            student_row[f"{code} - {subjects_dict[code]}"] = value
            if student_row["Name"] == "Not Found" and result["Name"] != "Not Found":
                student_row["Name"] = result["Name"]

        # GPA/Total
        gpa_data = get_gpa_result_total_percent(block, usn)
        total_data = get_total_and_percentage(block)
        student_row.update(
            {
                "SGPA": gpa_data["SGPA"],
                "CGPA": gpa_data["CGPA"],
                "Result": gpa_data["Result"],
                "Total Marks": clean_to_int(total_data["Total Marks"]),
                "Max Marks": clean_to_int(total_data["Max Marks"]),
                "Percentage": total_data["Percentage"],
            }
        )

        full_data.append(student_row)

    final_df = pd.DataFrame(full_data)

    # Sort by percentage (desc), add serial number, and order columns
    final_df = final_df.sort_values(by="Percentage", ascending=False)
    final_df.insert(0, "Sl. No", range(1, len(final_df) + 1))

    subject_columns = [f"{code} - {subjects_dict[code]}" for code in subjects_dict]
    final_columns = [
        "Sl. No",
        "Name",
        "USN",
        *subject_columns,
        "Result",
        "Total Marks",
        "Max Marks",
        "Percentage",
        "CGPA",
        "SGPA",
    ]
    final_df = final_df[final_columns]

    # Clean numeric columns
    final_df["Total Marks"] = pd.to_numeric(final_df["Total Marks"], errors="coerce").fillna(0).astype(int)
    final_df["Max Marks"] = pd.to_numeric(final_df["Max Marks"], errors="coerce").fillna(0).astype(int)
    final_df["Percentage"] = final_df["Percentage"].apply(
        lambda x: 0 if pd.isna(x) else round(x, 2) if isinstance(x, (float, int)) else 0.0
    )

    return final_df


def _process_single_file(input_path: str, mode: str) -> str:
    """Process one .xlsx file and write sibling _result.xlsx. Returns output path."""
    output_path = os.path.splitext(input_path)[0] + "_result.xlsx"
    logger.info(f"Reading: {input_path}")
    df = process_excel(input_path, mode=mode)
    logger.info(f"Writing: {output_path}")
    
    # Create Excel writer with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Results', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Results']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set minimum and maximum column widths
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply text wrapping to all cells
        from openpyxl.styles import Alignment
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Style the header row
        from openpyxl.styles import Font, PatternFill, Border, Side
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styling
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = thin_border
        
        # Apply borders to all data cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
        
        # Set print settings for better printing
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = False
        
        # Add alternating row colors for better readability
        from openpyxl.styles import PatternFill
        light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row_num in range(2, worksheet.max_row + 1):
            if row_num % 2 == 0:  # Even rows
                for cell in worksheet[row_num]:
                    if cell.fill.start_color.index == '00000000':  # Only if not already filled
                        cell.fill = light_fill
        
        # Set row height for better readability
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 25
    
    return output_path


def _process_path(path: str, mode: str) -> None:
    """Process a file or recursively process all .xlsx files inside a directory."""
    abs_path = os.path.abspath(path)
    if os.path.isdir(abs_path):
        logger.info(f"Scanning directory recursively: {abs_path}")
        processed = 0
        skipped = 0
        errors = 0
        for root, _dirs, files in os.walk(abs_path):
            for name in files:
                if not name.lower().endswith(".xlsx"):
                    continue
                if name.lower().endswith("_result.xlsx"):
                    skipped += 1
                    logger.debug(f"Skipping already processed: {os.path.join(root, name)}")
                    continue
                file_path = os.path.join(root, name)
                try:
                    _process_single_file(file_path, mode)
                    processed += 1
                except Exception as exc:
                    errors += 1
                    logger.error(f"Failed: {file_path} -> {exc}")
        logger.info(f"Done. Processed={processed}, Skipped={skipped}, Errors={errors}")
    else:
        if not abs_path.lower().endswith(".xlsx"):
            raise ValueError("Input must be a .xlsx file or a directory containing .xlsx files")
        if abs_path.lower().endswith("_result.xlsx"):
            logger.info("Input file already looks like a result file. Nothing to do.")
            return
        _process_single_file(abs_path, mode)


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Process a semester result Excel (with Sheet1 subjects, Sheet2 results) "
            "and produce a consolidated Excel next to the input with _result suffix. "
            "The script matches the logic used in result_analysis_app.py."
        )
    )
    parser.add_argument(
        "input",
        help=(
            "Path to an input .xlsx file OR a directory to scan recursively for .xlsx files"
        ),
    )
    parser.add_argument(
        "--mode",
        choices=["status", "marks", "combined"],
        default="combined",
        help=(
            "Subject columns mode: 'status' shows Pass/Fail only; 'marks' shows obtained/max only; "
            "'combined' shows obtained-status (e.g., '57-PASS'). Default is 'combined'."
        ),
    )
    args = parser.parse_args()

    _process_path(args.input, mode=args.mode)
    logger.info("All tasks completed.")


if __name__ == "__main__":
    main()